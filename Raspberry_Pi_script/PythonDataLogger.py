from openpyxl import Workbook, load_workbook
from threading import Thread
from statistics import mode
from scipy import stats as s
import urllib.request # importing lib for handles params and url
from urllib.request import urlopen
import http.client
import logging # importing lib for system base logging data
import os
import numpy
import time
import random
from threading import Thread
from datetime import datetime # import datatime for RPI time according to its RTC
import pygame, time
from pygame.locals import *
import RPi.GPIO as GPIO 
from RF24 import *
from RF24Network import *
##
##### radio setup for RPi B Rev2: CS0=Pin 24
radio = RF24(25,0);
network = RF24Network(radio)
####
octlit = lambda n:int(n,8)
####
this_node = octlit("00")
####
radio.begin()
radio.setPALevel(RF24_PA_MAX) # Power Amplifier
radio.setDataRate(RF24_250KBPS)
network.begin(90,this_node)
radio.printDetails()

class data_manager:
    def __init__(self):
        self.NodeID = 0
        self.Amp = 0.0
        self.Freq = 0.0
        self.Filename = "output/Motion_Records.xlsx"
        self.maxNode = 4
        self.buffer2 = numpy.zeros((self.maxNode,15),int)
        self.buffer3 = numpy.zeros((self.maxNode,15),int)
        self.buffer4 = numpy.zeros((self.maxNode,3,5),int)
        self.CP1 = numpy.zeros((self.maxNode,2),float)
        self.CP2 = numpy.zeros((self.maxNode,2),float)
        self.CP3 = numpy.zeros((self.maxNode,2),float)
        self.CP4 = numpy.zeros((self.maxNode,2),float)
        self.queue = []
        self.parameter = ""
        self.esc_pressed = 0
        self.mainStatus = 0
        self.thingSpeakStatus = 0
        self.internetstatus = 0
        if(not os.path.exists( self.Filename)):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "NODE_1"
            for idx in range(1,5):
                cellIndex = sheet.max_row
                sheet.cell(cellIndex,1).value = "TIMESTAMP"
                sheet.cell(cellIndex,2).value = "ACTION CODE"
                sheet.cell(cellIndex,3).value = "WEIGHT"
                sheet.cell(cellIndex,4).value = "ACTION 1"                
                sheet.cell(cellIndex,5).value = "CREDIT POINTS 1"
                sheet.cell(cellIndex,6).value = "PERIOD 1"
                sheet.cell(cellIndex,7).value = "ACTION 2"                
                sheet.cell(cellIndex,8).value = "CREDIT POINTS 2"
                sheet.cell(cellIndex,9).value = "PERIOD 2"
                sheet.cell(cellIndex,10).value = "ACTION 3"                
                sheet.cell(cellIndex,11).value = "CREDIT POINTS 3"
                sheet.cell(cellIndex,12).value = "PERIOD 3"
                sheet.cell(cellIndex,13).value = "ACTION 4"                
                sheet.cell(cellIndex,14).value = "CREDIT POINTS 4"
                sheet.cell(cellIndex,15).value = "PERIOD 4"
                cellIndex += 1
                sheet.cell(cellIndex,1).value = datetime.now().strftime("%I:%M:%S")
                sheet.cell(cellIndex,2).value = 0
                sheet.cell(cellIndex,3).value = 0
                sheet.cell(cellIndex,4).value = "NMD"                
                sheet.cell(cellIndex,5).value = 0
                sheet.cell(cellIndex,6).value = 0
                sheet.cell(cellIndex,7).value = "NMD"                
                sheet.cell(cellIndex,8).value = 0
                sheet.cell(cellIndex,9).value = 0
                sheet.cell(cellIndex,10).value = "NMD"                
                sheet.cell(cellIndex,11).value = 0
                sheet.cell(cellIndex,12).value = 0
                sheet.cell(cellIndex,13).value = "NMD"                
                sheet.cell(cellIndex,14).value = 0
                sheet.cell(cellIndex,15).value = 0
                workbook.save(self.Filename)
                workbook.create_sheet("NODE_" + str(idx+1))
                workbook.active = idx
                sheet = workbook.active
            workbook.close()
        else:
            workbook = load_workbook(self.Filename)
            idx = 0
            for idx in range(0,4):
                workbook.active = idx
                sheet = workbook.active
                cellIndex = sheet.max_row
                print(cellIndex)
                self.CP1[idx][0] = sheet.cell(cellIndex,5).value
                self.CP1[idx][1] = sheet.cell(cellIndex,6).value
                self.CP2[idx][0] = sheet.cell(cellIndex,8).value
                self.CP2[idx][1] = sheet.cell(cellIndex,9).value
                self.CP3[idx][0] = sheet.cell(cellIndex,11).value
                self.CP3[idx][1] = sheet.cell(cellIndex,12).value
                if(cellIndex > 5):
                    print("Node value:",idx)
                    self.CP4[idx][0] = sheet.cell(cellIndex-2,14).value
                    self.CP4[idx][1] = sheet.cell(cellIndex-2,15).value
                    self.buffer4[idx][0][0] = sheet.cell(cellIndex-4,2).value
                    self.buffer4[idx][0][1] = sheet.cell(cellIndex-3,2).value
                    self.buffer4[idx][0][2] = sheet.cell(cellIndex-2,2).value
                    self.buffer4[idx][0][3] = sheet.cell(cellIndex-1,2).value
                    self.buffer4[idx][0][4] = sheet.cell(cellIndex,2).value
                else:
                    self.CP4[idx][0] = 0
                    self.CP4[idx][1] = 0
                print(self.buffer4[idx][0])
##            print(self.CP1)
##            print(self.CP2)
##            print(self.CP3)
##            print(self.CP4)
            workbook.close()
        print("File: "+ self.Filename)

        self.database = []
        workbook = load_workbook("database/database.xlsx")        
        ##print(len(workbook.sheetnames))
        for idx in range(0,4):
            workbook.active = idx
            sheet = workbook.active
            templist = []
            templist.insert(0,("STATIC",0,0,"0-0",0))
            for itr in range(2,sheet.max_row + 1):
                templist.insert(itr,((sheet.cell(itr,1).value, sheet.cell(itr,2).value, sheet.cell(itr,3).value, sheet.cell(itr,4).value, sheet.cell(itr,5).value)))
            templist += [("NMD",0,0,"1000-1000",0)] 
            self.database.append(templist)
##        for m in range(0,4):
##            for i in range(0,len(self.database[m])):
##                #print(i,self.database[0][i],self.database[1][i],self.database[2][i],self.database[3][i])
##                print(i,self.database[m][i])
##            print("\n\n\n\n==========================================================================================")
        ##print(self.database[0],self.database[1],self.database[2],self.database[3])    
        workbook.close()
        
        ##==============================Syncing of Data===================================================##
        workbook = load_workbook("notSynced/syncbase.xlsx")
        sheet = workbook.active
        for index in range(1,sheet.max_row):
            self.queue.append(sheet.cell(index,1).value)
        workbook.remove(sheet)
        workbook.create_sheet("data",0)
        workbook.save("notSynced/syncbase.xlsx")
        workbook.close()
        print(self.queue)

    def send_to_Thingspeak(self):
        ##internetStatus = 0
        while(1):
            ##internetStatus = 0
            self.thingSpeakStatus = 1
            if(handler.esc_pressed == 1):
                break;
            if(len(handler.queue) > 0):
                try:
                    urlopen('https://www.google.com',timeout = 1)
                    self.internetstatus = 1
                    ##print("Internet status: Active")
                except Exception as Error:
                    self.internetstatus = 0
                    ##print("Internet status: Disconnected")
                if(self.internetstatus == 1):
                    headers = {"Content-typZZe": "application/x-www-form-urlencoded", "Accept": "text/plain"}
                    conn = http.client.HTTPConnection("api.thingspeak.com:80",timeout = 1)
                    
                    try:
                        conn.request("POST", "/update", self.parameter, headers) # send POST request to thingspeak.com
                        response = conn.getresponse() # get all response from POST request
                        data = response.read() # parse the results from response
                        ##print('data sent : {} Response :'.format(data),response)
                         # closing the connection of requesting
                        self.parameter = self.queue.pop(0)
                    except Exception as error:
                        print("connection failed : ", error)
        self.thingSpeakStatus = 0
        

    def receiveData(self):
        while(1):
            network.update()
            output = ""
            #print("Waiting for Data")
            #self.send_to_Thingspeak()
            #check_for_ESC()
            while network.available():
                header, payload = network.read(32)
                output = payload.decode("utf-8")
                #print("output: ",output)
                try:
                    if len(output) > 0:
                        splittedOutput = output.split(',')
                        ID = int(splittedOutput[0])
                        cnt = int(splittedOutput[1])
                        Frequency = float(splittedOutput[2])
                        Amplitude = float(splittedOutput[3])
                        return ID, cnt, Frequency, Amplitude
                except Exception as error:
                    print('Error occured while reading data. Error: {}'.format(error))
            return -1, 0, 0.0, 0.0

    def table_lookup(self, value1, value2, node):
        if(value2 < 1):
            print("STATIC", 0, 0, 0)
            return "STATIC", 0, 0, 0
  
        for index in range(0,len(self.database[node])):
            val_A = self.database[node][index][3]
            rangeArr = val_A.split("-")
            if float(rangeArr[0]) <= value2 <= float(rangeArr[1]): #parsing strings to float
                val_F =  self.database[node][index][2]#value 2 matched so checking value 1 now

                if isinstance(val_F, str): #if value 1 was entered in a string format then repeat like before
                    rangeArr = val_F.split("-")                
                    if float(rangeArr[0]) <= value1 <= float(rangeArr[1]):
                        ##print("In the interval===============================================")
                        print(self.database[node][index][0],self.database[node][index][1],self.database[node][index][4],index)
                        return  self.database[node][index][0],self.database[node][index][1],self.database[node][index][4],index#if matched then return action, weight, and R value

                elif value1 == val_F:#if value 1 was not a string/range and was a number
                    print(self.database[node][index][0],self.database[node][index][1],self.database[node][index][4],index)
                    return self.database[node][index][0],self.database[node][index][1],self.database[node][index][4],index
        print("NMD", 0, 0, len(self.database[node])-1)
        return "NMD", 0, 0, len(self.database[node])-1 #default return/if value is not contained in database

def calculate_e(W, R, t = 4, T = 600):#calculate Ep
    return (W*t*R)/T

def method2(node, actionIndex, itr):
    counter = 0
    method2Index = int((itr)%15)
    if(method2Index == 0):
        tempMode = int(s.mode(handler.buffer2[node])[0])
        for i in range(0,14):
            if(tempMode == handler.buffer2[node][i]):
                counter += 1

    handler.buffer2[node][method2Index] = actionIndex
    if(method2Index == 0):
        ##print(tempMode,handler.buffer2[node])
        if(float(counter)/15) < 0.7:
            handler.CP2[node][1] += 1
            return "NSA", handler.CP2[node][0], handler.CP2[node][1]
        else:
            #print(node,tempMode)
            tsp = calculate_e(handler.database[node][tempMode][1],handler.database[node][tempMode][4])
            handler.CP2[node][0] += tsp
            handler.CP2[node][1] += 1
            return handler.database[node][tempMode][0], handler.CP2[node][0], handler.CP2[node][1]
    return "---", handler.CP2[node][0], handler.CP2[node][1]

def method3(node, actionIndex, itr):
    if(itr >= 7):
        counter = 0
        method3Index = int((itr-7)%15)
        if(method3Index == 0):
            tempMode = int(s.mode(handler.buffer3[node])[0])
            for i in range(0,14):
                if(tempMode == handler.buffer3[node][i]):
                    counter += 1

        handler.buffer3[node][method3Index] = actionIndex
        if(method3Index == 0):
            ##print(tempMode,handler.buffer3[node])
            if(float(counter)/15) < 0.7:
                handler.CP3[node][1] += 1
                return "NSA", handler.CP3[node][0], handler.CP3[node][1]
            else:
                ##print(node,tempMode)
                tsp = calculate_e(handler.database[node][tempMode][1],handler.database[node][tempMode][4])
                handler.CP3[node][0] += tsp
                handler.CP3[node][1] += 1
                return handler.database[node][tempMode][0], handler.CP3[node][0], handler.CP3[node][1]
    return "---", handler.CP3[node][0], handler.CP3[node][1]

def method4(node,actionIndex):
    for i in range(0,4):
        handler.buffer4[node][0][i] = handler.buffer4[node][0][i+1]
    handler.buffer4[node][0][4] = actionIndex
    print(handler.buffer4[node][0])
    temparray =  numpy.zeros((3), int)
    output = handler.buffer4[node][0][2]
    for i in range(0,3):
        for j in range(0,3):
          temparray[j] =   handler.buffer4[node][0][j+i]
        MODE = int(s.mode(temparray)[0])
        error = 0
        for l in range(0,3):
            if(temparray[l] != MODE):
                error += 1
        #print("MODE , ERROR:",MODE,error)
        if error <= 1:
            #print("Position, Buffer:",i,handler.buffer4[node][0][2])
            output = MODE
        #print(temparray,output)
    tsp = calculate_e(handler.database[node][output][1],handler.database[node][output][4])
    handler.CP4[node][0] += tsp
    handler.CP4[node][1] += float(4)/60
    return handler.database[node][output][0], handler.CP4[node][0], handler.CP4[node][1]
    
def save_to_excel(node,weight,act,buf1,buf2,buf3,buf4,wb):
    try:
        
        wb.active = node
        sheet = wb.active
        cellIndex = sheet.max_row + 1
        sheet.cell(cellIndex,1).value = datetime.now().strftime("%d/%m/%Y, %I:%M:%S %p")
        sheet.cell(cellIndex,2).value = act
        sheet.cell(cellIndex,3).value = weight
        sheet.cell(cellIndex,4).value = buf1[0]               
        sheet.cell(cellIndex,5).value = buf1[1]  
        sheet.cell(cellIndex,6).value = buf1[2]  
        sheet.cell(cellIndex,7).value = buf2[0]                 
        sheet.cell(cellIndex,8).value = buf2[1]  
        sheet.cell(cellIndex,9).value = buf2[2]  
        sheet.cell(cellIndex,10).value = buf3[0]               
        sheet.cell(cellIndex,11).value = buf3[1]  
        sheet.cell(cellIndex,12).value = buf3[2]
        if(cellIndex > 4):
            sheet.cell(cellIndex-2,13).value = buf4[0]              
            sheet.cell(cellIndex-2,14).value = buf4[1]  
            sheet.cell(cellIndex-2,15).value = buf4[2]  
        
    finally:
        print("Sheet Updated....")


    
        


def main_program():
    iteration = numpy.zeros((handler.maxNode,1),int)
    method1Payload = ["ACTION",0,0]
    method2Payload = ["ACTION",0,0]
    method3Payload = ["ACTION",0,0]
    method4Payload = ["ACTION",0,0]
    ApiKeys = ["K6DJ233TUBX2WH70",
               "P4VYHYV4OWYE60IM",
               "QXVN0G6PAH7KDYOT",
               "USU3PO66I4OX5T8J",
               "NFZT4XR4GK5RGU84",
               "TS4X1YRPI4BSPXAC"]
    ApiKey = ""
    logging.basicConfig(format='%(asctime)s:%(levelname)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p',
                            filename= "LDS_Jeff_" + '30_05_2021.log', level=logging.DEBUG)
    wb = load_workbook(handler.Filename)

    while(1):
        handler.mainStatus = 1
##        nodeID = random.randint(1,2)
##        Freq = float(random.randint(2,10))/4
##        Amp = float(random.randint(10,60))
        nodeID,count, Freq, Amp = handler.receiveData()
        print("\n\n\nReceived:",nodeID,Freq, Amp)
        nodeID -= 1
        if(0 <= nodeID <= 3):
            ApiKey = ApiKeys[nodeID]
            action, W, R, actionIndex = handler.table_lookup(Freq, Amp, nodeID)
            curr_time = datetime.now() # getting real time datetime
            seconds = curr_time.strftime('%S') # reformat the seconds according to requirement
            logging.info('time={} , rpi time={} : {} , {} , {} , {a:.3f} , {p:.3f}'.format(seconds, curr_time, nodeID, datetime.now().strftime("%d-%m-%Y"),datetime.now().strftime("%I-%M-%S"), a=Amp,p=Freq))
            #actionIndex = random.randint(25,29)
            handler.CP1[nodeID][0] += calculate_e(W,R)
            handler.CP1[nodeID][1] += float(4)/60
            method1Payload = action, handler.CP1[nodeID][0], handler.CP1[nodeID][1]
            method2Payload = method2(nodeID,actionIndex,iteration[nodeID])
            method3Payload = method3(nodeID,actionIndex,iteration[nodeID])
            method4Payload = method4(nodeID,actionIndex)
            #print(method1Payload)
            #print(method2Payload)
            print(method1Payload,method2Payload,method3Payload, method4Payload)
            excelLogger = Thread(target=save_to_excel,args = (nodeID,W,actionIndex,method1Payload,method2Payload,method3Payload, method4Payload,wb))
            excelLogger.start()
            excelLogger.join()
            params = urllib.parse.urlencode(  # collect params for post request
                    {'field1': nodeID + 1, 'field2': datetime.now().strftime("%d/%m/%Y"), 'field3': datetime.now().strftime("%I:%M:%S"), 'field4': Amp, 'field5': Freq, 'field6':action, 'field7':handler.CP1[nodeID][1], 'field8':handler.CP1[nodeID][0],
                     'key': ApiKey})
            handler.queue.append(params)
            iteration[nodeID] += 1
            
            if(handler.internetstatus == 0):
                print("Internet is Disconnected. Packets Accumulated: ",len(handler.queue))
            else:
                print("Data sync is running in the background. Packets left: ",len(handler.queue))
        if(handler.esc_pressed == 1):
            wb.save(handler.Filename)
            break;
    time.sleep(2)
    handler.mainStatus = 0


if __name__ == '__main__':
    handler = data_manager()
    pygame.init()
    screen = pygame.display.set_mode((640, 480))
    pygame.display.set_caption('NRF FFT RECEIVE')
    pygame.mouse.set_visible(1)
    a = Thread(target = main_program)
    b = Thread(target = handler.send_to_Thingspeak)
#     a.daemon = True
#     b.daemon = True
#     c.daemon = True
    
    a.start()
    b.start()
    
    while(1):
        for event in pygame.event.get():
            if (event.type == KEYUP) or (event.type == KEYDOWN):
                if(event.key == 27):
                    pygame.quit()
                    handler.esc_pressed = 1
                    while(handler.mainStatus == 1) and ( handler.thingSpeakStatus == 1):
                        print("\nWaiting for other functions to close\n")
                        time.sleep(0.5)
                    workbook = load_workbook("notSynced/syncbase.xlsx")
                    sheet = workbook.active
                    cellPointer = 1
                    if(len(handler.queue) > 0):
                        data = handler.parameter
                    while(len(handler.queue) > 0):
                        ##print(cellPointer)
                        sheet.cell(cellPointer, 1).value = data
                        data = handler.queue.pop(0)
                        cellPointer = sheet.max_row + 1
                    workbook.save("notSynced/syncbase.xlsx")
                    workbook.close()
                    break
    print("DONE")

    
    









